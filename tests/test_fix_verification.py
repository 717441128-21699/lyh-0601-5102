import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import datetime, timedelta, date
from app.database import SessionLocal, engine, Base
from app.models import Customer, ChangeRequest, ApprovalRecord, Notification
from app.schemas import CustomerCreate, ChangeRequestCreate
from app.services import (
    customer_service,
    change_request_service,
    approval_service,
    approval_rule_service,
    risk_service,
    report_service,
    export_service
)
from app.scripts import init_approval_data
from app.services.utils import get_date_range


def run_fix_verification():
    print("=" * 60)
    print("修复验证测试 - 风控/催办/口径/日报")
    print("=" * 60)

    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    try:
        init_approval_data.init_default_approval_rules(db)

        print("\n【1】验证风控逻辑：第3次预警不拦截，第4次冻结")
        risk_customer = customer_service.create_customer(
            db,
            CustomerCreate(
                customer_code="FIX001",
                customer_name="风控修复测试客户",
                customer_level="NORMAL",
                contact_person="测试员",
                contact_phone="13600000001",
                contact_email="fix@test.com",
                address="测试地址",
                industry="测试行业",
                department="测试部门",
                order_manager="测试经理"
            ),
            "admin"
        )
        print(f"  ✓ 创建测试客户: {risk_customer.customer_name}")

        for i in range(3):
            result = change_request_service.submit_change_request(
                db,
                ChangeRequestCreate(
                    customer_code="FIX001",
                    change_type=f"TEST_{i}",
                    submitter=f"测试员{i}",
                    department="测试部门",
                    new_data={"contact_phone": f"1360000000{i + 2}"}
                )
            )
            status = result["request"]["status"]
            risk_triggered = result["risk_triggered"]
            approval_flow = result.get("approval_flow")
            print(f"  第{i+1}次提交: 状态={status}, 风控触发={risk_triggered}, 有审批流={approval_flow is not None}")

        status_3rd = change_request_service.get_change_request_detail(db, result["request"]["id"])
        risk_status = risk_service.get_customer_risk_status(db, risk_customer.id)
        print(f"\n  ✓ 3次后客户状态: {risk_status['status']}")
        print(f"  ✓ 3次后是否冻结: {risk_status['is_frozen']}")
        print(f"  ✓ 第3次申请状态: {status_3rd['status']} (应该是APPROVED，因为普通客户自动审批)")
        print(f"  ✓ 第3次是否有风控标记: {status_3rd.get('risk_triggered', False)}")
        assert risk_status["is_frozen"] == False, "第3次不应冻结"
        assert status_3rd["status"] == "APPROVED", "第3次应该正常审批通过"

        print("\n  第4次提交 - 应该触发冻结...")
        result_4th = change_request_service.submit_change_request(
            db,
            ChangeRequestCreate(
                customer_code="FIX001",
                change_type="TEST_FREEZE",
                submitter="测试员3",
                department="测试部门",
                new_data={"contact_phone": "13699999999"}
            )
        )
        print(f"  第4次提交: 状态={result_4th['request']['status']}, 风控触发={result_4th['risk_triggered']}")

        risk_status_4 = risk_service.get_customer_risk_status(db, risk_customer.id)
        print(f"  ✓ 4次后客户状态: {risk_status_4['status']}")
        print(f"  ✓ 4次后是否冻结: {risk_status_4['is_frozen']}")
        assert risk_status_4["is_frozen"] == True, "第4次应该冻结"
        assert result_4th["request"]["status"] == "RISK_HOLD", "第4次应该风控暂停"

        print("\n【2】验证风控解除后恢复流转")
        warnings = risk_service.get_risk_warnings(db, customer_id=risk_customer.id)
        warning_id = warnings["items"][0].id
        print(f"  ✓ 找到风控预警: {warning_id}")

        handled = risk_service.handle_risk_warning(
            db, warning_id, "风控管理员", "已核实，解除冻结", unfreeze=True
        )
        print(f"  ✓ 预警处理完成")

        risk_status_after = risk_service.get_customer_risk_status(db, risk_customer.id)
        print(f"  ✓ 解除后状态: {risk_status_after['status']}")
        print(f"  ✓ 解除后是否冻结: {risk_status_after['is_frozen']}")

        risk_hold_count = db.query(ChangeRequest).filter(
            ChangeRequest.customer_id == risk_customer.id,
            ChangeRequest.status == "RISK_HOLD"
        ).count()
        pending_count = db.query(ChangeRequest).filter(
            ChangeRequest.customer_id == risk_customer.id,
            ChangeRequest.status == "PENDING"
        ).count()
        print(f"  ✓ 风控暂停申请数: {risk_hold_count} (应为0)")
        print(f"  ✓ 待审批申请数: {pending_count} (应为1，因为第4次恢复后应该是PENDING)")
        assert risk_hold_count == 0, "解除后不应有风控暂停的申请"
        assert pending_count >= 1, "解除后应该有申请恢复为待审批"

        print("\n【3】验证统一日期口径")
        start_7d, end_7d = get_date_range(quick_range="7d")
        start_30d, end_30d = get_date_range(quick_range="30d")
        start_today, end_today = get_date_range(quick_range="today")

        print(f"  ✓ 7天范围: {start_7d.strftime('%Y-%m-%d %H:%M:%S')} ~ {end_7d.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"  ✓ 30天范围: {start_30d.strftime('%Y-%m-%d %H:%M:%S')} ~ {end_30d.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"  ✓ 今天范围: {start_today.strftime('%Y-%m-%d %H:%M:%S')} ~ {end_today.strftime('%Y-%m-%d %H:%M:%S')}")

        diff_days_7d = (end_7d - start_7d).days + 1
        diff_days_30d = (end_30d - start_30d).days + 1
        print(f"  ✓ 7天跨度: {diff_days_7d}天 (应该是7天)")
        print(f"  ✓ 30天跨度: {diff_days_30d}天 (应该是30天)")
        assert diff_days_7d == 7, "7天范围应该正好7天"
        assert diff_days_30d == 30, "30天范围应该正好30天"

        list_result = change_request_service.query_change_requests_quick(
            db, quick_range="today", page=1, page_size=100
        )
        print(f"\n  ✓ 列表查询今天: {list_result['total']} 条")

        from app.schemas import ChangeRequestQuery
        export_path = export_service.export_change_requests_excel(
            db,
            ChangeRequestQuery(page=1, page_size=100),
            quick_range="today"
        )
        print(f"  ✓ 导出今天数据: {os.path.basename(export_path)}")
        assert os.path.exists(export_path), "导出文件应该存在"

        print("\n【4】验证待办通知和超时催办")
        vip_customer = customer_service.create_customer(
            db,
            CustomerCreate(
                customer_code="FIX002",
                customer_name="待办测试VIP客户",
                customer_level="VIP",
                contact_person="VIP测试",
                contact_phone="13500000001",
                contact_email="vip@test.com",
                address="VIP地址",
                industry="互联网",
                department="销售二部",
                order_manager="VIP经理"
            ),
            "admin"
        )

        todo_result = change_request_service.submit_change_request(
            db,
            ChangeRequestCreate(
                customer_code="FIX002",
                change_type="TODO_TEST",
                submitter="提交员工",
                department="销售二部",
                new_data={"contact_phone": "13511111111"}
            ),
            priority="HIGH"
        )
        print(f"  ✓ 提交VIP变更申请: {todo_result['request']['request_no']}")
        print(f"  ✓ 命中规则: {todo_result['approval_flow']['matched_rule']}")
        print(f"  ✓ 当前状态: {todo_result['request']['status']}")

        notifications = db.query(Notification).filter(
            Notification.change_request_id == todo_result["request"]["id"]
        ).all()
        print(f"  ✓ 生成通知数: {len(notifications)}")
        if notifications:
            for n in notifications:
                print(f"    - {n.notification_type}: {n.title}")

        todo_before = approval_service.get_todo_stats(db, role="REGIONAL_MANAGER")
        print(f"\n  ✓ 区域经理待办统计: 总{todo_before['total_todo']}条, 高优{todo_before['high_priority_count']}条")

        print("\n【5】验证日报生成（前一天数据+7天/30天趋势+部门统计）")
        from datetime import timedelta
        test_date = date.today() - timedelta(days=1)

        report = report_service.generate_daily_report(db, report_date=test_date)
        print(f"  ✓ 日报生成日期: {report.report_date}")
        print(f"  ✓ 总申请数: {report.total_requests}")
        print(f"  ✓ 部门统计数: {len(report_service.json_to_dict(report.department_stats))}")

        report_detail = report_service.get_report_detail(db, test_date)
        dept_stats = report_detail["department_stats"]
        print(f"  ✓ 部门统计详情: {len(dept_stats)}个部门")
        for dept, stats in list(dept_stats.items())[:3]:
            print(f"    - {dept}: 申请{stats['total']}件, 通过率{stats['approval_rate']}%, "
                  f"同步成功率{stats['sync_success_rate']}%, "
                  f"平均时长{stats['avg_processing_hours']}h, "
                  f"超时{stats['overdue_count']}件")

        trend_7d = report_service.get_7day_trend(db, end_date=test_date)
        trend_30d = report_service.get_30day_trend(db, end_date=test_date)
        print(f"\n  ✓ 7天趋势数据: {len(trend_7d)}天")
        print(f"  ✓ 30天趋势数据: {len(trend_30d)}天")
        assert len(trend_7d) == 7, "7天趋势应该有7天数据"
        assert len(trend_30d) == 30, "30天趋势应该有30天数据"

        pdf_exists = os.path.exists(report.pdf_path) if report.pdf_path else False
        excel_exists = os.path.exists(report.excel_path) if report.excel_path else False
        print(f"\n  ✓ PDF报告存在: {pdf_exists}")
        print(f"  ✓ Excel报告存在: {excel_exists}")
        assert pdf_exists, "PDF报告应该存在"
        assert excel_exists, "Excel报告应该存在"

        from openpyxl import load_workbook
        wb = load_workbook(report.excel_path)
        sheet_names = wb.sheetnames
        print(f"  ✓ Excel工作表: {sheet_names}")
        assert "7日趋势" in sheet_names, "Excel应该包含7日趋势工作表"
        assert "30日趋势" in sheet_names, "Excel应该包含30日趋势工作表"
        assert "部门统计" in sheet_names, "Excel应该包含部门统计工作表"

        print("\n" + "=" * 60)
        print("所有修复验证通过！")
        print("=" * 60)

    finally:
        db.close()


if __name__ == "__main__":
    run_fix_verification()
