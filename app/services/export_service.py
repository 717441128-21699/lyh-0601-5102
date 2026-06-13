import os
import json
from io import BytesIO
from datetime import datetime, date, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import and_
from app.models import ChangeRequest, OperationLog, Customer
from app.schemas import ChangeRequestQuery
from app.services.utils import json_to_dict

EXPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)


def _get_date_range(quick_range: str = None, start_date: date = None, end_date: date = None):
    """统一日期范围计算口径"""
    if quick_range:
        today = date.today()
        if quick_range == "today":
            start_dt = datetime.combine(today, datetime.min.time())
            end_dt = datetime.combine(today, datetime.max.time())
        elif quick_range == "yesterday":
            yesterday = today - timedelta(days=1)
            start_dt = datetime.combine(yesterday, datetime.min.time())
            end_dt = datetime.combine(yesterday, datetime.max.time())
        elif quick_range == "7d":
            start_dt = datetime.combine(today - timedelta(days=6), datetime.min.time())
            end_dt = datetime.combine(today, datetime.max.time())
        elif quick_range == "30d":
            start_dt = datetime.combine(today - timedelta(days=29), datetime.min.time())
            end_dt = datetime.combine(today, datetime.max.time())
        else:
            start_dt, end_dt = None, None
    else:
        start_dt = datetime.combine(start_date, datetime.min.time()) if start_date else None
        end_dt = datetime.combine(end_date, datetime.max.time()) if end_date else None

    return start_dt, end_dt


def export_change_requests_excel(db: Session, query: ChangeRequestQuery, quick_range: str = None) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query_stmt = db.query(ChangeRequest)

    if query.customer_name:
        query_stmt = query_stmt.filter(
            ChangeRequest.customer_code.in_(
                db.query(Customer.customer_code).filter(
                    Customer.customer_name.like(f"%{query.customer_name}%")
                )
            )
        )

    start_dt, end_dt = _get_date_range(
        quick_range=quick_range,
        start_date=query.start_date,
        end_date=query.end_date
    )
    if start_dt:
        query_stmt = query_stmt.filter(ChangeRequest.created_at >= start_dt)
    if end_dt:
        query_stmt = query_stmt.filter(ChangeRequest.created_at <= end_dt)

    if query.status:
        query_stmt = query_stmt.filter(ChangeRequest.status == query.status)

    if query.department:
        query_stmt = query_stmt.filter(ChangeRequest.department == query.department)

    items = query_stmt.order_by(ChangeRequest.created_at.desc()).all()

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"change_requests_{timestamp}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "变更申请明细"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        "申请单号", "客户编码", "客户名称", "变更类型", "提交人",
        "所属部门", "状态", "审批人", "审批时间", "同步状态",
        "同步时间", "变更字段", "创建时间"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    customer_cache = {}

    for row_idx, item in enumerate(items, start=2):
        if item.customer_code not in customer_cache:
            customer = db.query(Customer).filter(Customer.customer_code == item.customer_code).first()
            customer_cache[item.customer_code] = customer.customer_name if customer else ""

        diff_data = json_to_dict(item.diff_data)
        diff_fields = ", ".join([d["field"] for d in diff_data]) if diff_data else ""

        row_data = [
            item.request_no,
            item.customer_code,
            customer_cache[item.customer_code],
            item.change_type,
            item.submitter,
            item.department,
            _get_status_text(item.status),
            item.approver or "",
            item.approved_at.strftime("%Y-%m-%d %H:%M:%S") if item.approved_at else "",
            _get_sync_status_text(item.sync_status),
            item.synced_at.strftime("%Y-%m-%d %H:%M:%S") if item.synced_at else "",
            diff_fields,
            item.created_at.strftime("%Y-%m-%d %H:%M:%S")
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    column_widths = [20, 15, 20, 12, 12, 15, 10, 12, 18, 12, 18, 30, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = width

    wb.save(filepath)
    return filepath


def export_operation_logs_excel(db: Session, start_date: date = None, end_date: date = None,
                                 operator: str = None, operation_type: str = None,
                                 quick_range: str = None) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = db.query(OperationLog)

    start_dt, end_dt = _get_date_range(
        quick_range=quick_range,
        start_date=start_date,
        end_date=end_date
    )
    if start_dt:
        query = query.filter(OperationLog.created_at >= start_dt)
    if end_dt:
        query = query.filter(OperationLog.created_at <= end_dt)

    if operator:
        query = query.filter(OperationLog.operator == operator)

    if operation_type:
        query = query.filter(OperationLog.operation_type == operation_type)

    items = query.order_by(OperationLog.created_at.desc()).all()

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"operation_logs_{timestamp}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "操作日志"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="#70AD47", end_color="#70AD47", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["日志ID", "操作类型", "操作人", "目标类型", "目标ID", "操作详情", "IP地址", "操作时间"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, item in enumerate(items, start=2):
        row_data = [
            item.id,
            item.operation_type,
            item.operator,
            item.target_type,
            item.target_id or "",
            item.detail,
            item.ip_address or "",
            item.created_at.strftime("%Y-%m-%d %H:%M:%S")
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    column_widths = [10, 18, 15, 15, 10, 50, 15, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    wb.save(filepath)
    return filepath


def _get_status_text(status: str) -> str:
    status_map = {
        "PENDING": "待审批",
        "APPROVED": "已通过",
        "REJECTED": "已驳回",
        "RISK_HOLD": "风控暂停",
        "CANCELLED": "已取消"
    }
    return status_map.get(status, status)


def _get_sync_status_text(status: str) -> str:
    status_map = {
        "PENDING": "待同步",
        "SUCCESS": "全部成功",
        "PARTIAL": "部分成功",
        "FAILED": "全部失败",
        "BLOCKED": "已阻塞",
        "CANCELLED": "已取消"
    }
    return status_map.get(status, status)


def batch_export_change_details(db: Session, request_ids: list) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    items = db.query(ChangeRequest).filter(ChangeRequest.id.in_(request_ids)).all()

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"batch_export_{timestamp}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "变更汇总"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["申请单号", "客户名称", "变更类型", "提交人", "状态", "同步状态", "创建时间"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    customer_cache = {}

    ws2 = wb.create_sheet("变更明细")
    detail_headers = ["申请单号", "字段名", "旧值", "新值"]
    for col, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="#70AD47", end_color="#70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    detail_row = 2

    for row_idx, item in enumerate(items, start=2):
        if item.customer_code not in customer_cache:
            customer = db.query(Customer).filter(Customer.customer_code == item.customer_code).first()
            customer_cache[item.customer_code] = customer.customer_name if customer else ""

        row_data = [
            item.request_no,
            customer_cache[item.customer_code],
            item.change_type,
            item.submitter,
            _get_status_text(item.status),
            _get_sync_status_text(item.sync_status),
            item.created_at.strftime("%Y-%m-%d %H:%M:%S")
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

        diff_data = json_to_dict(item.diff_data)
        for diff in diff_data:
            ws2.cell(row=detail_row, column=1, value=item.request_no).border = thin_border
            ws2.cell(row=detail_row, column=2, value=diff["field"]).border = thin_border
            ws2.cell(row=detail_row, column=3, value=str(diff.get("old_value", ""))).border = thin_border
            ws2.cell(row=detail_row, column=4, value=str(diff.get("new_value", ""))).border = thin_border
            detail_row += 1

    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 10
    ws1.column_dimensions['F'].width = 12
    ws1.column_dimensions['G'].width = 20

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 30
    ws2.column_dimensions['D'].width = 30

    wb.save(filepath)
    return filepath
