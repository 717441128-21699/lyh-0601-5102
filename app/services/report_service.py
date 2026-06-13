import os
import json
from datetime import datetime, date, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import func, and_
from app.models import ChangeRequest, DailyReport, RiskWarning, Customer
from app.services.utils import dict_to_json, json_to_dict, log_operation

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)


def generate_daily_report(db: Session, report_date: date = None) -> DailyReport:
    if report_date is None:
        report_date = date.today() - timedelta(days=1)

    existing = db.query(DailyReport).filter(DailyReport.report_date == report_date).first()
    if existing:
        return existing

    start_of_day = datetime.combine(report_date, datetime.min.time())
    end_of_day = datetime.combine(report_date, datetime.max.time())

    day_requests = db.query(ChangeRequest).filter(
        and_(
            ChangeRequest.created_at >= start_of_day,
            ChangeRequest.created_at <= end_of_day
        )
    ).all()

    total_requests = len(day_requests)
    approved_count = sum(1 for r in day_requests if r.status == "APPROVED")
    approval_rate = (approved_count / total_requests * 100) if total_requests > 0 else 0.0

    approved_requests = [r for r in day_requests if r.status == "APPROVED"]
    sync_success_count = sum(1 for r in approved_requests if r.sync_status == "SUCCESS")
    sync_success_rate = (sync_success_count / len(approved_requests) * 100) if len(approved_requests) > 0 else 0.0

    processing_hours = []
    for r in day_requests:
        if r.approved_at and r.created_at:
            delta = r.approved_at - r.created_at
            processing_hours.append(delta.total_seconds() / 3600)
    avg_processing_hours = sum(processing_hours) / len(processing_hours) if processing_hours else 0.0

    overdue_count = sum(1 for r in day_requests if r.is_overdue)

    department_stats = _calculate_department_stats(day_requests)
    change_type_stats = {}
    for r in day_requests:
        ctype = r.change_type or "其他"
        if ctype not in change_type_stats:
            change_type_stats[ctype] = 0
        change_type_stats[ctype] += 1

    risk_count = db.query(RiskWarning).filter(
        and_(
            RiskWarning.created_at >= start_of_day,
            RiskWarning.created_at <= end_of_day
        )
    ).count()

    report = DailyReport(
        report_date=report_date,
        total_requests=total_requests,
        approved_count=approved_count,
        approval_rate=round(approval_rate, 2),
        sync_success_count=sync_success_count,
        sync_success_rate=round(sync_success_rate, 2),
        avg_processing_hours=round(avg_processing_hours, 2),
        department_stats=dict_to_json(department_stats),
        change_type_stats=dict_to_json(change_type_stats),
        risk_warning_count=risk_count,
        overdue_count=overdue_count
    )

    db.add(report)
    db.flush()

    pdf_path = generate_pdf_report(report)
    excel_path = generate_excel_report(report)

    report.pdf_path = pdf_path
    report.excel_path = excel_path

    log_operation(
        db,
        operation_type="GENERATE_REPORT",
        operator="SYSTEM",
        target_type="DAILY_REPORT",
        target_id=report.id,
        detail=f"生成日报: {report_date}"
    )

    db.commit()
    db.refresh(report)
    return report


def _calculate_department_stats(requests: list) -> dict:
    """计算各部门详细统计：申请量、通过率、同步成功率、平均处理时长"""
    dept_stats = {}
    for r in requests:
        dept = r.department or "未分配"
        if dept not in dept_stats:
            dept_stats[dept] = {
                "total": 0,
                "approved": 0,
                "rejected": 0,
                "pending": 0,
                "approval_rate": 0.0,
                "sync_success": 0,
                "sync_total": 0,
                "sync_success_rate": 0.0,
                "processing_hours": [],
                "avg_processing_hours": 0.0,
                "overdue_count": 0
            }

        stats = dept_stats[dept]
        stats["total"] += 1

        if r.status == "APPROVED":
            stats["approved"] += 1
            stats["sync_total"] += 1
            if r.sync_status == "SUCCESS":
                stats["sync_success"] += 1

            if r.approved_at and r.created_at:
                delta = r.approved_at - r.created_at
                stats["processing_hours"].append(delta.total_seconds() / 3600)
        elif r.status == "REJECTED":
            stats["rejected"] += 1
        else:
            stats["pending"] += 1

        if r.is_overdue:
            stats["overdue_count"] += 1

    for dept, stats in dept_stats.items():
        stats["approval_rate"] = round(
            stats["approved"] / stats["total"] * 100, 2
        ) if stats["total"] > 0 else 0.0
        stats["sync_success_rate"] = round(
            stats["sync_success"] / stats["sync_total"] * 100, 2
        ) if stats["sync_total"] > 0 else 0.0
        stats["avg_processing_hours"] = round(
            sum(stats["processing_hours"]) / len(stats["processing_hours"]), 2
        ) if stats["processing_hours"] else 0.0
        del stats["processing_hours"]

    return dept_stats


def get_7day_trend(db: Session, end_date: date = None) -> list:
    return get_trend_data(db, days=7, end_date=end_date)


def get_30day_trend(db: Session, end_date: date = None) -> list:
    return get_trend_data(db, days=30, end_date=end_date)


def get_trend_data(db: Session, days: int = 7, end_date: date = None) -> list:
    """获取N天趋势数据"""
    if end_date is None:
        end_date = date.today() - timedelta(days=1)

    trend_data = []
    for i in range(days - 1, -1, -1):
        d = end_date - timedelta(days=i)
        report = db.query(DailyReport).filter(DailyReport.report_date == d).first()
        if report:
            dept_stats = json_to_dict(report.department_stats)
            trend_data.append({
                "date": d.strftime("%Y-%m-%d"),
                "total_requests": report.total_requests,
                "approved_count": report.approved_count,
                "approval_rate": report.approval_rate,
                "sync_success_count": report.sync_success_count,
                "sync_success_rate": report.sync_success_rate,
                "avg_processing_hours": report.avg_processing_hours,
                "risk_warning_count": report.risk_warning_count,
                "overdue_count": report.overdue_count,
                "department_count": len(dept_stats)
            })
        else:
            trend_data.append({
                "date": d.strftime("%Y-%m-%d"),
                "total_requests": 0,
                "approved_count": 0,
                "approval_rate": 0,
                "sync_success_count": 0,
                "sync_success_rate": 0,
                "avg_processing_hours": 0,
                "risk_warning_count": 0,
                "overdue_count": 0,
                "department_count": 0
            })

    return trend_data


def generate_pdf_report(report: DailyReport) -> str:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    import platform
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    system_name = platform.system()
    if system_name == "Windows":
        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei']
    elif system_name == "Darwin":
        plt.rcParams['font.sans-serif'] = ['Arial Unicode MS']
    else:
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False

    pdf_path = os.path.join(REPORTS_DIR, f"daily_report_{report.report_date}.pdf")

    doc = SimpleDocTemplate(pdf_path, pagesize=A4,
                            leftMargin=2 * cm, rightMargin=2 * cm,
                            topMargin=2 * cm, bottomMargin=2 * cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18, spaceAfter=20)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=14, spaceAfter=10)
    normal_style = styles['Normal']

    story = []

    story.append(Paragraph("客户主数据变更管理日报", title_style))
    story.append(Paragraph(f"报告日期: {report.report_date}", subtitle_style))
    story.append(Spacer(1, 0.5 * cm))

    summary_data = [
        ["指标", "数值"],
        ["变更申请总数", str(report.total_requests)],
        ["审批通过数", str(report.approved_count)],
        ["审批通过率", f"{report.approval_rate}%"],
        ["同步成功数", str(report.sync_success_count)],
        ["同步成功率", f"{report.sync_success_rate}%"],
        ["平均处理时长(小时)", str(report.avg_processing_hours)],
        ["风控预警数", str(report.risk_warning_count)],
        ["超时申请数", str(report.overdue_count)]
    ]

    summary_table = Table(summary_data, colWidths=[8 * cm, 6 * cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F2F2F2')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#D9D9D9'))
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.8 * cm))

    chart_7d_path = _generate_trend_chart_png(report.report_date, 7)
    if chart_7d_path:
        story.append(Paragraph("7日趋势图", subtitle_style))
        img = Image(chart_7d_path, width=16 * cm, height=8 * cm)
        story.append(img)
        story.append(Spacer(1, 0.5 * cm))

    chart_30d_path = _generate_trend_chart_png(report.report_date, 30)
    if chart_30d_path:
        story.append(Paragraph("30日趋势图", subtitle_style))
        img = Image(chart_30d_path, width=16 * cm, height=8 * cm)
        story.append(img)
        story.append(Spacer(1, 0.5 * cm))

    dept_stats = json_to_dict(report.department_stats)
    if dept_stats:
        story.append(Paragraph("各部门详细统计", subtitle_style))
        dept_data = [["部门", "申请数", "通过数", "通过率", "同步成功率", "平均时长(h)", "超时数"]]
        for dept, stats in dept_stats.items():
            dept_data.append([
                dept,
                str(stats["total"]),
                str(stats["approved"]),
                f"{stats['approval_rate']}%",
                f"{stats['sync_success_rate']}%",
                str(stats["avg_processing_hours"]),
                str(stats["overdue_count"])
            ])
        dept_table = Table(dept_data, colWidths=[3.5 * cm, 2 * cm, 2 * cm, 2 * cm, 2.5 * cm, 2.5 * cm, 1.5 * cm])
        dept_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#70AD47')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#D9D9D9'))
        ]))
        story.append(dept_table)
        story.append(Spacer(1, 0.5 * cm))

    doc.build(story)
    return pdf_path


def _generate_trend_chart_png(report_date: date, days: int = 7) -> str:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from sqlalchemy.orm import Session
    from app.database import SessionLocal
    import platform

    system_name = platform.system()
    if system_name == "Windows":
        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei']
    elif system_name == "Darwin":
        plt.rcParams['font.sans-serif'] = ['Arial Unicode MS']
    else:
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False

    db = SessionLocal()
    try:
        trend = get_trend_data(db, days=days, end_date=report_date)
    finally:
        db.close()

    if not trend:
        return None

    dates = [d["date"][5:] for d in trend]
    totals = [d["total_requests"] for d in trend]
    approved = [d["approved_count"] for d in trend]
    approval_rates = [d["approval_rate"] for d in trend]

    chart_path = os.path.join(REPORTS_DIR, f"trend_{days}d_{report_date}.png")

    fig, ax1 = plt.subplots(figsize=(12, 5))

    bar_width = 0.35
    x = range(len(dates))

    ax1.bar([i - bar_width / 2 for i in x], totals, width=bar_width,
            label='申请总数', color='#4472C4', alpha=0.8)
    ax1.bar([i + bar_width / 2 for i in x], approved, width=bar_width,
            label='通过数', color='#70AD47', alpha=0.8)

    ax1.set_xlabel('日期')
    ax1.set_ylabel('申请数量')
    ax1.set_title(f'{days}日申请趋势')
    ax1.set_xticks(x)
    ax1.set_xticklabels(dates)
    ax1.legend(loc='upper left')

    ax2 = ax1.twinx()
    ax2.plot(x, approval_rates, color='#ED7D31', marker='o', linewidth=2, label='通过率(%)')
    ax2.set_ylabel('通过率(%)')
    ax2.legend(loc='upper right')

    plt.tight_layout()
    plt.savefig(chart_path, dpi=100, bbox_inches='tight')
    plt.close()

    return chart_path


def generate_excel_report(report: DailyReport) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference

    excel_path = os.path.join(REPORTS_DIR, f"daily_report_{report.report_date}.xlsx")

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "汇总"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws1.merge_cells('A1:B1')
    ws1['A1'] = f"客户主数据变更管理日报 - {report.report_date}"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].alignment = center_align

    summary_data = [
        ("指标", "数值"),
        ("变更申请总数", report.total_requests),
        ("审批通过数", report.approved_count),
        ("审批通过率(%)", report.approval_rate),
        ("同步成功数", report.sync_success_count),
        ("同步成功率(%)", report.sync_success_rate),
        ("平均处理时长(小时)", report.avg_processing_hours),
        ("风控预警数", report.risk_warning_count),
        ("超时申请数", report.overdue_count)
    ]

    for row_idx, (key, value) in enumerate(summary_data, start=3):
        ws1.cell(row=row_idx, column=1, value=key)
        ws1.cell(row=row_idx, column=2, value=value)
        if row_idx == 3:
            for col in range(1, 3):
                cell = ws1.cell(row=row_idx, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
        else:
            for col in range(1, 3):
                cell = ws1.cell(row=row_idx, column=col)
                cell.border = thin_border
                cell.alignment = center_align

    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20

    ws2 = wb.create_sheet("部门统计")
    dept_stats = json_to_dict(report.department_stats)

    headers = ["部门", "申请数", "通过数", "驳回数", "待审批",
               "通过率(%)", "同步成功数", "同步成功率(%)", "平均时长(h)", "超时数"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, (dept, stats) in enumerate(dept_stats.items(), start=2):
        row_data = [
            dept,
            stats["total"],
            stats["approved"],
            stats["rejected"],
            stats["pending"],
            stats["approval_rate"],
            stats["sync_success"],
            stats["sync_success_rate"],
            stats["avg_processing_hours"],
            stats["overdue_count"]
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = center_align

    for col_idx, width in enumerate([20, 10, 10, 10, 10, 12, 12, 12, 12, 10], 1):
        ws2.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)].width = width

    if dept_stats:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "各部门申请统计"
        chart.y_axis.title = '数量'
        chart.x_axis.title = '部门'

        data = Reference(ws2, min_col=2, min_row=1, max_row=len(dept_stats) + 1, max_col=5)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=len(dept_stats) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 20
        chart.height = 10
        ws2.add_chart(chart, "L2")

    ws3 = wb.create_sheet("变更类型统计")
    type_stats = json_to_dict(report.change_type_stats)

    ws3.cell(row=1, column=1, value="变更类型").font = header_font
    ws3.cell(row=1, column=1).fill = header_fill
    ws3.cell(row=1, column=1).alignment = center_align
    ws3.cell(row=1, column=2, value="数量").font = header_font
    ws3.cell(row=1, column=2).fill = header_fill
    ws3.cell(row=1, column=2).alignment = center_align

    for row_idx, (ctype, count) in enumerate(type_stats.items(), start=2):
        ws3.cell(row=row_idx, column=1, value=ctype).border = thin_border
        ws3.cell(row=row_idx, column=2, value=count).border = thin_border

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 15

    from app.database import SessionLocal
    db_session = SessionLocal()
    try:
        trend_7d = get_trend_data(db_session, days=7, end_date=report.report_date)
        trend_30d = get_trend_data(db_session, days=30, end_date=report.report_date)
    finally:
        db_session.close()

    ws4 = wb.create_sheet("7日趋势")
    trend_headers = ["日期", "申请总数", "通过数", "通过率(%)", "同步成功数", "同步成功率(%)", "平均时长(h)", "风控预警数", "超时数"]
    for col, header in enumerate(trend_headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, item in enumerate(trend_7d, start=2):
        row_data = [
            item["date"],
            item["total_requests"],
            item["approved_count"],
            item["approval_rate"],
            item["sync_success_count"],
            item["sync_success_rate"],
            item["avg_processing_hours"],
            item["risk_warning_count"],
            item["overdue_count"]
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = center_align

    for col_idx, width in enumerate([12, 10, 10, 12, 12, 14, 12, 12, 10], 1):
        ws4.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)].width = width

    if trend_7d:
        chart7 = BarChart()
        chart7.type = "col"
        chart7.style = 10
        chart7.title = "7日申请趋势"
        chart7.y_axis.title = '数量'
        chart7.x_axis.title = '日期'

        data = Reference(ws4, min_col=2, min_row=1, max_row=len(trend_7d) + 1, max_col=3)
        cats = Reference(ws4, min_col=1, min_row=2, max_row=len(trend_7d) + 1)
        chart7.add_data(data, titles_from_data=True)
        chart7.set_categories(cats)
        chart7.width = 20
        chart7.height = 10
        ws4.add_chart(chart7, "K2")

    ws5 = wb.create_sheet("30日趋势")
    for col, header in enumerate(trend_headers, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, item in enumerate(trend_30d, start=2):
        row_data = [
            item["date"],
            item["total_requests"],
            item["approved_count"],
            item["approval_rate"],
            item["sync_success_count"],
            item["sync_success_rate"],
            item["avg_processing_hours"],
            item["risk_warning_count"],
            item["overdue_count"]
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws5.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = center_align

    for col_idx, width in enumerate([12, 10, 10, 12, 12, 14, 12, 12, 10], 1):
        ws5.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)].width = width

    if trend_30d:
        chart30 = BarChart()
        chart30.type = "col"
        chart30.style = 10
        chart30.title = "30日申请趋势"
        chart30.y_axis.title = '数量'
        chart30.x_axis.title = '日期'

        data = Reference(ws5, min_col=2, min_row=1, max_row=len(trend_30d) + 1, max_col=3)
        cats = Reference(ws5, min_col=1, min_row=2, max_row=len(trend_30d) + 1)
        chart30.add_data(data, titles_from_data=True)
        chart30.set_categories(cats)
        chart30.width = 30
        chart30.height = 10
        ws5.add_chart(chart30, "K2")

    wb.save(excel_path)
    return excel_path


def get_reports_list(db: Session, page: int = 1, page_size: int = 20) -> dict:
    query = db.query(DailyReport)
    total = query.count()
    items = query.order_by(DailyReport.report_date.desc()) \
        .offset((page - 1) * page_size) \
        .limit(page_size) \
        .all()

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": items
    }


def get_report_by_date(db: Session, report_date: date) -> DailyReport:
    return db.query(DailyReport).filter(DailyReport.report_date == report_date).first()


def get_report_detail(db: Session, report_date: date) -> dict:
    """获取日报详情，包含部门统计结构化数据"""
    report = get_report_by_date(db, report_date)
    if not report:
        return None

    return {
        "id": report.id,
        "report_date": report.report_date,
        "total_requests": report.total_requests,
        "approved_count": report.approved_count,
        "approval_rate": report.approval_rate,
        "sync_success_count": report.sync_success_count,
        "sync_success_rate": report.sync_success_rate,
        "avg_processing_hours": report.avg_processing_hours,
        "risk_warning_count": report.risk_warning_count,
        "overdue_count": report.overdue_count,
        "department_stats": json_to_dict(report.department_stats),
        "change_type_stats": json_to_dict(report.change_type_stats),
        "pdf_path": report.pdf_path,
        "excel_path": report.excel_path,
        "created_at": report.created_at
    }
